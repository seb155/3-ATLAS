"""
Tests for Package Export System

Tests:
- Package CRUD operations
- Template export (IN-P040, CA-P040)
- Excel file generation
- Asset association
"""

import io
from datetime import datetime

import pytest
from openpyxl import load_workbook

from app.models.auth import Client, Project
from app.models.cables import Cable
from app.models.models import Asset, AssetType
from app.models.packages import Package
from app.services.template_service import TemplateService


@pytest.fixture
def test_package_project(db_session):
    """Create test project for package tests"""
    # Create client first
    client = Client(id="test-client-pkg", name="Test Package Client")
    db_session.add(client)

    project = Project(
        id="test-project-pkg",
        name="Test Package Project",
        client_id="test-client-pkg",
    )
    db_session.add(project)
    db_session.commit()
    return project


@pytest.fixture
def test_instruments(db_session, test_package_project):
    """Create test instruments for IN-P040 export"""
    instruments = [
        Asset(
            tag="FIT-210-001",
            description="Flow Transmitter - Feed Water",
            type="INSTRUMENT",
            project_id="test-project-pkg",
            area="210",
            discipline="Instrumentation",
            electrical={
                "voltage": "24VDC",
                "power_supply": "UPS-210-01",
                "signal_type": "4-20mA",
                "io_points": 1,
                "panel": "PLC-210-01",
            },
            process={"service": "Feed Water Flow"},
        ),
        Asset(
            tag="PIT-210-002",
            description="Pressure Transmitter - Discharge",
            type="INSTRUMENT",
            project_id="test-project-pkg",
            area="210",
            discipline="Instrumentation",
            electrical={
                "voltage": "24VDC",
                "power_supply": "UPS-210-01",
                "signal_type": "4-20mA",
                "io_points": 1,
                "panel": "PLC-210-01",
            },
            process={"service": "Discharge Pressure"},
        ),
        Asset(
            tag="TIT-210-003",
            description="Temperature Transmitter - Inlet",
            type="INSTRUMENT",
            project_id="test-project-pkg",
            area="210",
            discipline="Instrumentation",
            electrical={
                "voltage": "24VDC",
                "power_supply": "UPS-210-01",
                "signal_type": "RTD",
                "io_points": 1,
                "panel": "PLC-210-01",
            },
            process={"service": "Inlet Temperature"},
        ),
    ]
    for inst in instruments:
        db_session.add(inst)
    db_session.commit()
    return instruments


@pytest.fixture
def test_cables(db_session, test_package_project):
    """Create test cables for CA-P040 export"""
    cables = [
        Cable(
            tag="PWR-210-001",
            project_id="test-project-pkg",
            cable_type="POWER",
            conductor_size="4x25mm2",
            length_meters=45.0,
            description="Power cable from MCC to Motor",
        ),
        Cable(
            tag="SIG-210-001",
            project_id="test-project-pkg",
            cable_type="SIGNAL",
            conductor_size="2x1.5mm2",
            length_meters=120.0,
            description="Signal cable from PLC to FIT",
        ),
    ]
    for cable in cables:
        db_session.add(cable)
    db_session.commit()
    return cables


@pytest.fixture
def test_package(db_session, test_package_project, test_instruments):
    """Create test package with instruments"""
    package = Package(
        name="Area 210 Instrumentation",
        description="Instrumentation package for Area 210 - Process Plant",
        package_type="IN-P040",
        project_id="test-project-pkg",
        package_metadata={
            "discipline": "Instrumentation",
            "area": "210",
            "revision": "A",
            "prepared_by": "Test Engineer",
        },
    )
    db_session.add(package)
    db_session.flush()

    # Add instruments to package
    for inst in test_instruments:
        package.assets.append(inst)

    db_session.commit()
    return package


class TestPackageCRUD:
    """Test Package CRUD operations via API"""

    def test_create_package(self, client, test_package_project):
        """Test creating a new package"""
        from app.api.deps import get_current_active_user
        from app.main import app

        app.dependency_overrides[get_current_active_user] = lambda: {
            "username": "testuser"
        }

        response = client.post(
            "/api/v1/packages",
            headers={"X-Project-ID": "test-project-pkg"},
            json={
                "name": "Test Package",
                "description": "Test package description",
                "package_type": "IN-P040",
                "package_metadata": {"discipline": "Electrical"},
            },
        )

        assert response.status_code in [200, 201]  # Accept both OK and Created
        data = response.json()
        assert data["name"] == "Test Package"
        assert data["package_type"] == "IN-P040"
        print("✅ test_create_package passed")

    def test_list_packages(self, client, test_package_project, test_package):
        """Test listing packages"""
        from app.api.deps import get_current_active_user
        from app.main import app

        app.dependency_overrides[get_current_active_user] = lambda: {
            "username": "testuser"
        }

        response = client.get(
            "/api/v1/packages", headers={"X-Project-ID": "test-project-pkg"}
        )

        assert response.status_code == 200
        data = response.json()
        packages = data.get("packages", data) if isinstance(data, dict) else data
        assert len(packages) >= 1
        assert any(p["name"] == "Area 210 Instrumentation" for p in packages)
        print("✅ test_list_packages passed")

    def test_get_package_details(self, client, test_package):
        """Test getting package details"""
        from app.api.deps import get_current_active_user
        from app.main import app

        app.dependency_overrides[get_current_active_user] = lambda: {
            "username": "testuser"
        }

        response = client.get(
            f"/api/v1/packages/{test_package.id}",
            headers={"X-Project-ID": "test-project-pkg"},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["name"] == "Area 210 Instrumentation"
        assert data["package_type"] == "IN-P040"
        assert "assets" in data or "asset_count" in data
        print("✅ test_get_package_details passed")


class TestTemplateExport:
    """Test Template Export functionality"""

    def test_export_instrument_index(self, db_session, test_package, test_instruments):
        """Test IN-P040 Instrument Index export"""
        service = TemplateService(db_session)

        # Export to Excel
        result = service.export_package(
            package_id=test_package.id, template_type="IN-P040", format="xlsx"
        )

        assert result.success is True
        assert result.file_name.endswith(".xlsx")
        assert result.file_data is not None

        # Load workbook and verify content
        wb = load_workbook(io.BytesIO(result.file_data))
        ws = wb.active

        # Check headers exist (row 4 typically has column headers)
        headers_found = False
        for row in ws.iter_rows(min_row=1, max_row=10):
            cell_values = [str(cell.value).lower() for cell in row if cell.value]
            if "tag" in cell_values or "tag number" in cell_values:
                headers_found = True
                break

        assert headers_found, "Headers not found in exported Excel"

        # Check data rows - should have at least 3 instruments
        data_rows = 0
        for row in ws.iter_rows(min_row=5):
            # Check if row has tag number (first few columns)
            if any(
                cell.value
                and isinstance(cell.value, str)
                and ("FIT-210" in cell.value or "PIT-210" in cell.value or "TIT-210" in cell.value)
                for cell in list(row)[:5]
            ):
                data_rows += 1

        assert data_rows >= 3, f"Expected 3+ instruments, found {data_rows}"
        print("✅ test_export_instrument_index passed")

    def test_export_cable_schedule(self, db_session, test_package_project, test_cables):
        """Test CA-P040 Cable Schedule export"""
        # Create assets for cable connections
        motor = Asset(
            tag="MTR-210-001",
            description="Motor for pump",
            type="MOTOR",
            project_id="test-project-pkg",
            properties={"hp": 50, "voltage": "480V"},
        )
        mcc = Asset(
            tag="MCC-210-001",
            description="Motor Control Center",
            type="MCC",
            project_id="test-project-pkg",
            properties={"type": "MCC"},
        )
        db_session.add_all([motor, mcc])
        db_session.flush()

        # Create cable package with assets
        package = Package(
            name="Area 210 Power Cables",
            description="Power cable schedule for Area 210",
            package_type="CA-P040",
            project_id="test-project-pkg",
        )
        db_session.add(package)
        db_session.flush()

        # Assign assets to package
        motor.package_id = package.id
        mcc.package_id = package.id

        # Update cables to connect our assets
        test_cables[0].from_asset_id = mcc.id
        test_cables[0].to_asset_id = motor.id
        db_session.commit()

        service = TemplateService(db_session)

        # Export to Excel
        result = service.export_package(
            package_id=package.id, template_type="CA-P040", format="xlsx"
        )

        assert result.success is True
        assert result.file_name.endswith(".xlsx")
        assert result.file_data is not None

        # Load workbook and verify content
        wb = load_workbook(io.BytesIO(result.file_data))
        ws = wb.active

        # Check headers exist
        headers_found = False
        for row in ws.iter_rows(min_row=1, max_row=10):
            cell_values = [str(cell.value).lower() for cell in row if cell.value]
            if "cable" in " ".join(cell_values):
                headers_found = True
                break

        assert headers_found, "Headers not found in cable schedule"
        print("✅ test_export_cable_schedule passed")

    def test_export_via_api(self, client, test_package):
        """Test package export via API endpoint"""
        from app.api.deps import get_current_active_user
        from app.main import app

        app.dependency_overrides[get_current_active_user] = lambda: {
            "username": "testuser"
        }

        response = client.get(
            f"/api/v1/packages/{test_package.id}/export",
            headers={"X-Project-ID": "test-project-pkg"},
            params={"template_type": "IN-P040", "format": "xlsx"},
        )

        assert response.status_code == 200
        assert (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            in response.headers["content-type"]
        )
        assert len(response.content) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(io.BytesIO(response.content))
        assert wb.active is not None
        print("✅ test_export_via_api passed")

    def test_export_preview(self, client, test_package):
        """Test export preview endpoint"""
        from app.api.deps import get_current_active_user
        from app.main import app

        app.dependency_overrides[get_current_active_user] = lambda: {
            "username": "testuser"
        }

        response = client.get(
            f"/api/v1/packages/{test_package.id}/export/preview",
            headers={"X-Project-ID": "test-project-pkg"},
            params={"template_type": "IN-P040"},
        )

        assert response.status_code == 200
        data = response.json()
        assert "assets" in data or "data" in data
        assert "package" in data or "metadata" in data
        print("✅ test_export_preview passed")


class TestPackageAssets:
    """Test Package Asset Management"""

    def test_add_asset_to_package(self, client, db_session, test_package, test_package_project):
        """Test adding asset to package"""
        from app.api.deps import get_current_active_user
        from app.main import app

        app.dependency_overrides[get_current_active_user] = lambda: {
            "username": "testuser"
        }

        # Create a new asset
        new_asset = Asset(
            tag="LIT-210-004",
            description="Level Transmitter",
            type="INSTRUMENT",
            project_id="test-project-pkg",
        )
        db_session.add(new_asset)
        db_session.commit()

        response = client.post(
            f"/api/v1/packages/{test_package.id}/assets/{new_asset.id}",
            headers={"X-Project-ID": "test-project-pkg"},
        )

        assert response.status_code == 204
        print("✅ test_add_asset_to_package passed")

    def test_list_package_assets(self, client, test_package):
        """Test listing assets in a package"""
        from app.api.deps import get_current_active_user
        from app.main import app

        app.dependency_overrides[get_current_active_user] = lambda: {
            "username": "testuser"
        }

        response = client.get(
            f"/api/v1/packages/{test_package.id}/assets",
            headers={"X-Project-ID": "test-project-pkg"},
        )

        assert response.status_code == 200
        data = response.json()
        assets = data.get("assets", data) if isinstance(data, dict) else data
        assert isinstance(assets, list)
        assert len(assets) >= 3  # Our test instruments
        print("✅ test_list_package_assets passed")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
