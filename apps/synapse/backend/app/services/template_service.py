"""
Template Service

Generates Excel/PDF deliverables from templates using:
- Jinja2 for template processing
- openpyxl for Excel generation
- WeasyPrint for PDF generation (future)

Templates supported:
- IN-P040: Instrument Index (Panel instrumentation list)
- CA-P040: Cable Schedule (Power & signal cables)
- Package custom templates

Design based on: .dev/design/2025-11-28-whiteboard-session.md
"""

import io
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_
from sqlalchemy.orm import Session

from app.models.cables import Cable
from app.models.models import Asset
from app.models.packages import Package


@dataclass
class TemplateContext:
    """Context data for template rendering."""

    package: Package
    assets: list[Asset]
    cables: list[Cable] | None = None
    project_info: dict | None = None
    metadata: dict | None = None


@dataclass
class ExportResult:
    """Result of template export."""

    success: bool
    file_name: str
    file_data: bytes | None = None
    mime_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    error: str | None = None


class TemplateService:
    """
    Service for generating deliverables from templates.

    Features:
    - Excel templates with openpyxl
    - Jinja2 template processing
    - Auto-formatting and styling
    - Multi-sheet support
    """

    def __init__(self, db: Session, templates_dir: str | None = None):
        self.db = db
        self.templates_dir = Path(templates_dir or "app/templates")
        self._setup_jinja()

    def _setup_jinja(self):
        """Initialize Jinja2 environment."""
        self.jinja_env = Environment(
            loader=FileSystemLoader(str(self.templates_dir)),
            autoescape=select_autoescape(["html", "xml"]),
        )

        # Add custom filters
        self.jinja_env.filters["format_date"] = self._format_date
        self.jinja_env.filters["format_number"] = self._format_number

    @staticmethod
    def _format_date(value: datetime | None, format: str = "%Y-%m-%d") -> str:
        """Format datetime for templates."""
        if not value:
            return ""
        return value.strftime(format)

    @staticmethod
    def _format_number(value: float | None, decimals: int = 2) -> str:
        """Format number for templates."""
        if value is None:
            return ""
        return f"{value:.{decimals}f}"

    # ==========================================================================
    # MAIN EXPORT METHODS
    # ==========================================================================

    def export_package(
        self,
        package_id: str,
        template_type: str,
        format: str = "xlsx",
    ) -> ExportResult:
        """
        Export a package using specified template.

        Args:
            package_id: Package to export
            template_type: Template type (IN-P040, CA-P040, etc.)
            format: Output format (xlsx, pdf)

        Returns:
            ExportResult with file data
        """
        # Load package with assets
        package = self.db.query(Package).filter(Package.id == package_id).first()
        if not package:
            return ExportResult(
                success=False, file_name="", error=f"Package {package_id} not found"
            )

        # Load assets
        assets = (
            self.db.query(Asset)
            .filter(Asset.package_id == package_id)
            .order_by(Asset.tag)
            .all()
        )

        if not assets:
            return ExportResult(
                success=False,
                file_name="",
                error=f"Package {package.name} has no assets",
            )

        # Build context
        context = TemplateContext(
            package=package,
            assets=assets,
            project_info=self._get_project_info(package.project_id),
            metadata=self._get_metadata(),
        )

        # Route to appropriate template generator
        if template_type == "IN-P040":
            return self._export_instrument_index(context, format)
        elif template_type == "CA-P040":
            # Load cables for this package
            asset_ids = [a.id for a in assets]
            try:
                cables = (
                    self.db.query(Cable)
                    .filter(
                        and_(
                            Cable.from_asset_id.in_(asset_ids),
                            Cable.to_asset_id.in_(asset_ids),
                        )
                    )
                    .all()
                )
                context.cables = cables
            except Exception:
                # Cable table may not exist or have different schema
                context.cables = []
            return self._export_cable_schedule(context, format)
        else:
            return ExportResult(
                success=False,
                file_name="",
                error=f"Unknown template type: {template_type}",
            )

    # ==========================================================================
    # IN-P040: INSTRUMENT INDEX
    # ==========================================================================

    def _export_instrument_index(
        self, context: TemplateContext, format: str
    ) -> ExportResult:
        """
        Generate IN-P040 Instrument Index.

        Columns:
        - Tag Number
        - Service Description
        - Type
        - Location
        - Power Supply
        - Signal Type
        - IO Points
        - Panel
        - Remarks
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Instrument Index"

            # Header
            self._write_header(ws, context)

            # Column headers
            headers = [
                "Item",
                "Tag Number",
                "Service Description",
                "Type",
                "Location",
                "Power Supply",
                "Signal Type",
                "IO Points",
                "Panel",
                "Remarks",
            ]
            self._write_column_headers(ws, 6, headers)

            # Data rows
            row = 7
            for idx, asset in enumerate(context.assets, start=1):
                props = asset.properties or {}
                ws.cell(row, 1, idx)
                ws.cell(row, 2, asset.tag)
                ws.cell(row, 3, asset.description or props.get("description", ""))
                ws.cell(row, 4, asset.type)
                ws.cell(row, 5, props.get("location", ""))
                ws.cell(row, 6, props.get("power_supply", ""))
                ws.cell(row, 7, props.get("signal_type", ""))
                ws.cell(row, 8, props.get("io_points", ""))
                ws.cell(row, 9, props.get("panel", ""))
                ws.cell(row, 10, props.get("remarks", ""))
                row += 1

            # Auto-size columns
            self._auto_size_columns(ws)

            # Footer
            self._write_footer(ws, row + 1, context)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            file_name = f"{context.package.name}_IN-P040_{datetime.now().strftime('%Y%m%d')}.xlsx"

            return ExportResult(
                success=True,
                file_name=file_name,
                file_data=output.getvalue(),
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            return ExportResult(
                success=False, file_name="", error=f"Export failed: {str(e)}"
            )

    # ==========================================================================
    # CA-P040: CABLE SCHEDULE
    # ==========================================================================

    def _export_cable_schedule(
        self, context: TemplateContext, format: str
    ) -> ExportResult:
        """
        Generate CA-P040 Cable Schedule.

        Columns:
        - Cable Number
        - From (Equipment)
        - To (Equipment)
        - Cable Type
        - Core/Size
        - Length (m)
        - Routing
        - Tray/Duct
        - Termination From
        - Termination To
        - Remarks
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cable Schedule"

            # Header
            self._write_header(ws, context)

            # Column headers
            headers = [
                "Item",
                "Cable Number",
                "From Equipment",
                "To Equipment",
                "Cable Type",
                "Core/Size",
                "Length (m)",
                "Routing",
                "Tray/Duct",
                "Term. From",
                "Term. To",
                "Remarks",
            ]
            self._write_column_headers(ws, 6, headers)

            # Data rows
            if not context.cables:
                # No cables, show message
                ws.cell(7, 2, "No cables found for this package")
                row = 8
            else:
                row = 7
                for idx, cable in enumerate(context.cables, start=1):
                    from_asset = (
                        self.db.query(Asset)
                        .filter(Asset.id == cable.from_asset_id)
                        .first()
                    )
                    to_asset = (
                        self.db.query(Asset)
                        .filter(Asset.id == cable.to_asset_id)
                        .first()
                    )

                    ws.cell(row, 1, idx)
                    ws.cell(row, 2, cable.cable_number)
                    ws.cell(row, 3, from_asset.tag if from_asset else "")
                    ws.cell(row, 4, to_asset.tag if to_asset else "")
                    ws.cell(row, 5, cable.cable_type or "")
                    ws.cell(
                        row,
                        6,
                        f"{cable.cores}C x {cable.size_mm2}mm²"
                        if cable.cores and cable.size_mm2
                        else "",
                    )
                    ws.cell(row, 7, cable.length_m or "")
                    ws.cell(row, 8, cable.routing or "")
                    ws.cell(row, 9, cable.tray or "")
                    ws.cell(row, 10, cable.from_termination or "")
                    ws.cell(row, 11, cable.to_termination or "")
                    ws.cell(row, 12, cable.remarks or "")
                    row += 1

            # Auto-size columns
            self._auto_size_columns(ws)

            # Footer
            self._write_footer(ws, row + 1, context)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            file_name = f"{context.package.name}_CA-P040_{datetime.now().strftime('%Y%m%d')}.xlsx"

            return ExportResult(
                success=True,
                file_name=file_name,
                file_data=output.getvalue(),
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            return ExportResult(
                success=False, file_name="", error=f"Export failed: {str(e)}"
            )

    # ==========================================================================
    # FORMATTING HELPERS
    # ==========================================================================

    def _write_header(self, ws, context: TemplateContext, start_row: int = 1):
        """Write document header (project info, package, etc.)."""
        # Title
        ws.merge_cells(f"A{start_row}:J{start_row}")
        title_cell = ws.cell(start_row, 1, "SYNAPSE - MBSE Platform")
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Project info
        ws.merge_cells(f"A{start_row + 1}:J{start_row + 1}")
        project_cell = ws.cell(
            start_row + 1,
            1,
            f"Project: {context.project_info.get('name', 'Unknown') if context.project_info else 'Unknown'}",
        )
        project_cell.font = Font(size=12)
        project_cell.alignment = Alignment(horizontal="center")

        # Package info
        ws.merge_cells(f"A{start_row + 2}:J{start_row + 2}")
        package_cell = ws.cell(
            start_row + 2, 1, f"Package: {context.package.name}"
        )
        package_cell.font = Font(size=11, bold=True)
        package_cell.alignment = Alignment(horizontal="center")

        # Date
        ws.merge_cells(f"A{start_row + 3}:J{start_row + 3}")
        date_cell = ws.cell(
            start_row + 3, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )
        date_cell.font = Font(size=9)
        date_cell.alignment = Alignment(horizontal="center")

    def _write_column_headers(
        self, ws, row: int, headers: list[str], freeze: bool = True
    ):
        """Write and format column headers."""
        # Header style
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Freeze panes at header row
        if freeze:
            ws.freeze_panes = ws.cell(row + 1, 1)

    def _write_footer(self, ws, row: int, context: TemplateContext):
        """Write document footer."""
        ws.merge_cells(f"A{row}:J{row}")
        footer_cell = ws.cell(
            row,
            1,
            f"Total Items: {len(context.assets)} | Generated by SYNAPSE MBSE Platform",
        )
        footer_cell.font = Font(size=9, italic=True)
        footer_cell.alignment = Alignment(horizontal="center")

    def _auto_size_columns(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-size columns based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    # ==========================================================================
    # DATA HELPERS
    # ==========================================================================

    def _get_project_info(self, project_id: str) -> dict:
        """Get project information for header."""
        from app.models.auth import Project

        project = self.db.query(Project).filter(Project.id == project_id).first()
        if not project:
            return {"name": "Unknown Project"}

        return {
            "name": project.name,
            "description": project.description or "",
            "created_at": project.created_at,
        }

    def _get_metadata(self) -> dict:
        """Get metadata for footer."""
        return {
            "generator": "SYNAPSE MBSE Platform",
            "version": "0.2.3",
            "timestamp": datetime.now(),
        }
